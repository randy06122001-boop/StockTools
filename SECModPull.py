import requests
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter


def get_sec_data(ticker, email):
    """
    Pulls XBRL financial data from the SEC EDGAR API and formats it into a
    financial model layout with derived metrics (Gross Profit, margins).

    Returns a pivoted DataFrame with metrics as rows and periods as
    chronologically-ordered columns, or None on failure.
    """

    headers = {'User-Agent': email}

    # ---- 1. Resolve CIK for the ticker ----
    try:
        resp = requests.get(
            "https://www.sec.gov/files/company_tickers.json",
            headers=headers,
            timeout=15,
        )
        resp.raise_for_status()
        tickers_json = resp.json()
    except requests.RequestException as e:
        print(f"[ERROR] Failed to fetch ticker list from SEC: {e}")
        return None

    cik_str = None
    for value in tickers_json.values():
        if value['ticker'] == ticker.upper():
            cik_str = str(value['cik_str']).zfill(10)
            break

    if not cik_str:
        print(f"[ERROR] Could not find CIK for '{ticker}'")
        return None

    # ---- 2. Pull Company Facts (all historical XBRL data) ----
    try:
        url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik_str}.json"
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        facts = resp.json()['facts']['us-gaap']
    except requests.RequestException as e:
        print(f"[ERROR] Failed to fetch XBRL facts for {ticker}: {e}")
        return None
    except KeyError:
        print(f"[ERROR] No US-GAAP data found for {ticker}")
        return None

    # ---- 3. GAAP concept mapping (broad coverage) ----
    gaap_mapping = {
        'Revenue': [
            'Revenues',
            'SalesRevenueNet',
            'RevenueFromContractWithCustomerExcludingAssessedTax',
            'RevenueFromContractWithCustomerIncludingAssessedTax',
            'SalesRevenueGoodsNet',
        ],
        'COGS': [
            'CostOfGoodsAndServicesSold',
            'CostOfRevenue',
            'CostOfGoodsSold',
        ],
        'R&D': [
            'ResearchAndDevelopmentExpense',
        ],
        'SG&A': [
            'SellingGeneralAndAdministrativeExpense',
            'SellingAndMarketingExpense',   # fallback if SGA isn't reported
        ],
        'Total Operating Expenses': [
            'OperatingExpenses',
        ],
        'Operating Income': [
            'OperatingIncomeLoss',
        ],
        'Interest Expense': [
            'InterestExpense',
            'InterestExpenseNet',
        ],
        'Net Income': [
            'NetIncomeLoss',
        ],
        'EPS (Basic)': [
            'EarningsPerShareBasic',
        ],
        'EPS (Diluted)': [
            'EarningsPerShareDiluted',
        ],
        'Total Assets': [
            'Assets',
        ],
        'Total Liabilities': [
            'Liabilities',
        ],
        'Stockholders Equity': [
            'StockholdersEquity',
            'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest',
        ],
    }

    # Metrics that should NOT be converted to millions
    per_share_metrics = {'EPS (Basic)', 'EPS (Diluted)'}

    model_data = []

    # ---- 4. Extract & filter data points ----
    for row_name, gaap_tags in gaap_mapping.items():
        for tag in gaap_tags:
            if tag not in facts:
                continue
            units_key = list(facts[tag]['units'].keys())[0]
            data_points = facts[tag]['units'][units_key]

            for point in data_points:
                form = point.get('form', '')
                if form not in ('10-Q', '10-K', '8-K'):
                    continue

                # Safe period label
                fp = point.get('fp', '')
                fy = point.get('fy', '')
                if fp == 'FY':
                    period = str(fy)
                elif fp.startswith('Q') and len(fp) >= 2:
                    period = f"{fp} {str(fy)[-2:]}"
                else:
                    continue  # skip unparseable period codes

                value = point['val']
                if row_name not in per_share_metrics:
                    value = value / 1_000_000  # convert to millions

                model_data.append({
                    'Metric': row_name,
                    'Period': period,
                    'Value': value,
                    'End_Date': point['end'],
                    'Filed': point.get('filed', ''),
                })
            break  # use first matching tag per metric

    if not model_data:
        print(f"[WARNING] No XBRL data points extracted for {ticker}")
        return None

    # ---- 5. Build & pivot the DataFrame ----
    df = pd.DataFrame(model_data)

    # Keep the most-recently-filed value for each metric+period
    df = df.sort_values(['End_Date', 'Filed']).drop_duplicates(
        subset=['Metric', 'Period'], keep='last',
    )

    pivot_df = df.pivot(index='Metric', columns='Period', values='Value')

    # Sort columns chronologically using the latest End_Date per period
    period_order = (
        df.drop_duplicates(subset='Period', keep='last')
        .sort_values('End_Date')['Period']
        .tolist()
    )
    pivot_df = pivot_df[[p for p in period_order if p in pivot_df.columns]]

    # Reorder rows to match the mapping order
    row_order = [m for m in gaap_mapping.keys() if m in pivot_df.index]
    pivot_df = pivot_df.reindex(row_order)

    # ---- 6. Derived metrics ----
    if 'Revenue' in pivot_df.index and 'COGS' in pivot_df.index:
        pivot_df.loc['Gross Profit'] = pivot_df.loc['Revenue'] - pivot_df.loc['COGS']
        pivot_df.loc['Gross Margin %'] = (
            pivot_df.loc['Gross Profit'] / pivot_df.loc['Revenue'] * 100
        ).round(2)

    if 'Revenue' in pivot_df.index and 'Operating Income' in pivot_df.index:
        pivot_df.loc['Operating Margin %'] = (
            pivot_df.loc['Operating Income'] / pivot_df.loc['Revenue'] * 100
        ).round(2)

    if 'Revenue' in pivot_df.index and 'Net Income' in pivot_df.index:
        pivot_df.loc['Net Margin %'] = (
            pivot_df.loc['Net Income'] / pivot_df.loc['Revenue'] * 100
        ).round(2)

    # Final row ordering (insert derived rows after their natural position)
    final_order = []
    for m in row_order:
        final_order.append(m)
        if m == 'COGS':
            for derived in ('Gross Profit', 'Gross Margin %'):
                if derived in pivot_df.index:
                    final_order.append(derived)
        if m == 'Operating Income':
            if 'Operating Margin %' in pivot_df.index:
                final_order.append('Operating Margin %')
        if m == 'Net Income':
            if 'Net Margin %' in pivot_df.index:
                final_order.append('Net Margin %')
    # Append any remaining rows not yet in the order
    for m in pivot_df.index:
        if m not in final_order:
            final_order.append(m)
    pivot_df = pivot_df.reindex(final_order)

    return pivot_df


def export_to_excel(df, filepath, ticker):
    """
    Exports the financial model DataFrame to a formatted Excel workbook.
    """
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Financials')
        ws = writer.sheets['Financials']

        # ---- Styling constants ----
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        metric_font = Font(bold=True, size=11)
        number_fmt = '#,##0.00'
        pct_fmt = '0.00"%"'
        eps_fmt = '0.00'

        pct_rows = {'Gross Margin %', 'Operating Margin %', 'Net Margin %'}
        eps_rows = {'EPS (Basic)', 'EPS (Diluted)'}

        # ---- Format header row (row 1) ----
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # ---- Format data rows ----
        for row_idx in range(2, ws.max_row + 1):
            metric_cell = ws.cell(row=row_idx, column=1)
            metric_cell.font = metric_font
            metric_name = metric_cell.value

            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='right')

                if metric_name in pct_rows:
                    cell.number_format = pct_fmt
                elif metric_name in eps_rows:
                    cell.number_format = eps_fmt
                else:
                    cell.number_format = number_fmt

        # ---- Auto-size columns ----
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

        # Title row above headers
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=f"{ticker} — Historical Financial Model (USD in millions)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color='1F4E79')


# ---------------------------------------------------------------------------
# Execution
# ---------------------------------------------------------------------------
if __name__ == '__main__':

    # SEC requires a valid User-Agent (your email)
    MY_EMAIL = ""
    TICKER = "MP"

    if not MY_EMAIL:
        print("[ERROR] Set MY_EMAIL to a valid email address — SEC requires a User-Agent.")
    else:
        print(f"Pulling SEC data for {TICKER}...")
        model_df = get_sec_data(TICKER, MY_EMAIL)

        if model_df is not None:
            out_path = f"{TICKER}_Historical_Model.xlsx"
            export_to_excel(model_df, out_path, TICKER)
            print(f"Exported to {out_path}")
        else:
            print("[ERROR] No data returned — check the ticker and try again.")